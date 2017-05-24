#!/usr/bin/python
# -*- coding:utf-8 -*-
###############################################################################
#Author: arvon
#Email: yafeng2011@126.com
#Blog: http://blog.arvon.top/
#Date: 2017-05-23
#Filename: write_jumpserver_host.py
#Revision: 1.0
#License: GPL
#Description: use ansible get host list then use xlrd module write excel
#Notes:
###############################################################################
import os
import openpyxl
#vars
port='22'
host_group='xxx'
aws_access_id='xxx'
aws_secret_id='xxx'
#create_file_script='create-host.yml'
server_file_name='./running_ec2_list.txt'
dest_filename = 'asset_cn_dev.xlsx'

##functions
def create_server_file():
#    subprocess.call("sh (%s)" %(env_script),shell=True)
#    subprocess.call("ansible-playbook %(script_file)" %(create_file_script),shell=True)
    os.environ['AWS_ACCESS_KEY_ID'] = aws_access_id
    os.environ['AWS_SECRET_ACCESS_KEY'] = aws_secret_id
    os.system('ansible-playbook create-host.yml')
def write_jumpserver_excel():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Assets'
    ws1.append(['IP地址','端口号','主机名','管理账号','用户名','密码','主机组'])
    with open (server_file_name, 'r') as f1:
        server_num = len(open(server_file_name,'rU').readlines())
        row = int(2)
        for eachline in f1:
            server_info=eachline.split()
#            print server_info[0],server_info[1]
            private_ip=server_info[0]
            tag_name=server_info[1]
#            for row in range(2,2+server_num):
            ws1.cell(column=1,row=row,value=private_ip)
            ws1.cell(column=2,row=row,value=port)
            ws1.cell(column=3,row=row,value=tag_name)
            ws1.cell(column=4,row=row,value='默认')
            ws1.cell(column=5,row=row,value='')
            ws1.cell(column=6, row=row, value='')
            ws1.cell(column=7, row=row, value=host_group)
            wb.save(filename=dest_filename)
            row=row+1

if __name__=='__main__':
#    create_server_file(server_file_name)
    write_jumpserver_excel()
